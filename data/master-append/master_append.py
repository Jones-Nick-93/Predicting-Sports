"""
master_append.py
Append engine for tabular reports that accumulate into a single Excel master.

Solves a specific problem: a scheduled report runs daily, and the master
workbook must stay correct whether the job runs once, runs twice by
accident, or backfills a day that was missed. Different report types need
different reconciliation behavior, so the mode is explicit per call rather
than inferred.

Usage, replacing a df.to_excel(...) call:

    from master_append import append_to_master

    append_to_master(
        df=report_df,
        master_path="./output/report_master.xlsx",
        sheet_name="Records",
        dedup_keys=["Record_ID", "Line_ID"],
        rerun_mode="dedup",
        backup_dir="./output/backups",
    )

rerun_mode:
    "replace_snapshot" : delete existing rows carrying today's snapshot
                         value, then append. Re-running the same day is
                         safe. Use for daily state snapshots where a row
                         legitimately recurs across days.
    "dedup"            : append only rows whose dedup_keys combination is
                         not already present in the master, and drop
                         duplicates within the incoming batch. Use for
                         event data with a natural key.
    "append"           : blind append, no protection.
    "replace_all"      : wipe the sheet and write only the current pull.
                         Use when the query already returns full state
                         every run and there is nothing to accumulate.

Snapshot_Date is stored as a real Excel date value rather than text, so
downstream relative date filters ("last 7 days") work against it directly.
A value that does not parse as a date is stored unchanged.
"""

import os
import shutil
import tempfile
import warnings
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

VALID_MODES = ("replace_snapshot", "dedup", "append", "replace_all")


def _normalize_snapshot(value):
    """
    Returns (stored_value, compare_key).

    stored_value: a real datetime when the input looks like a date,
        otherwise the original value unchanged.
    compare_key: a normalized 'YYYY-MM-DD' string used for equality checks
        regardless of whether a value was previously stored as text, date,
        or Timestamp. This is what makes the engine tolerant of masters
        written by earlier versions.
    """
    if isinstance(value, datetime):
        return value, value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
        return dt, dt.strftime("%Y-%m-%d")
    if isinstance(value, pd.Timestamp):
        dt = value.to_pydatetime()
        return dt, dt.strftime("%Y-%m-%d")
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.notna(parsed):
        dt = parsed.to_pydatetime()
        return dt, dt.strftime("%Y-%m-%d")
    return value, str(value)


def append_to_master(
    df,
    master_path,
    sheet_name,
    dedup_keys=None,
    snapshot_col="Snapshot_Date",
    snapshot_value=None,
    rerun_mode="replace_snapshot",
    keep_snapshots=None,
    keep_backup=True,
    backup_dir=None,
    backup_keep=10,
    verbose=True,
):
    if df is None or len(df) == 0:
        if verbose:
            print(f"[{sheet_name}] No rows to append. Master untouched.")
        return None

    if rerun_mode not in VALID_MODES:
        raise ValueError(f"rerun_mode must be one of {VALID_MODES}")
    if rerun_mode == "dedup" and not dedup_keys:
        raise ValueError("rerun_mode='dedup' requires dedup_keys")

    raw_snapshot = snapshot_value if snapshot_value is not None else date.today()
    stored_snapshot, compare_key = _normalize_snapshot(raw_snapshot)

    new = df.copy()
    new.columns = [str(c).strip() for c in new.columns]
    if snapshot_col not in new.columns:
        new.insert(0, snapshot_col, stored_snapshot)
    else:
        new[snapshot_col] = new[snapshot_col].fillna(stored_snapshot)

    old = pd.DataFrame()
    if os.path.exists(master_path):
        try:
            old = pd.read_excel(master_path, sheet_name=sheet_name, dtype=object)
            old.columns = [str(c).strip() for c in old.columns]
        except ValueError:
            pass  # workbook exists, sheet does not yet

    added, removed, intra = len(new), 0, 0

    if rerun_mode == "replace_all":
        removed = len(old)
        old = pd.DataFrame()

    elif rerun_mode == "dedup":
        missing = [k for k in dedup_keys if k not in new.columns]
        if missing:
            raise KeyError(f"dedup_keys not found in incoming data: {missing}")
        # Fix: a single pull can contain the same key twice. Collapse the
        # incoming batch before comparing it against the master, otherwise
        # duplicates enter silently.
        before = len(new)
        new = new.drop_duplicates(subset=dedup_keys, keep="first")
        intra = before - len(new)
        if not old.empty:
            missing_old = [k for k in dedup_keys if k not in old.columns]
            if missing_old:
                raise KeyError(f"dedup_keys not found in existing master: {missing_old}")
            old_keys = set(map(tuple, old[dedup_keys].astype(str).values))
            keep = ~new[dedup_keys].astype(str).apply(tuple, axis=1).isin(old_keys)
            new = new.loc[keep]
        added = len(new)

    elif rerun_mode == "replace_snapshot" and not old.empty:
        if snapshot_col in old.columns:
            old_keys = old[snapshot_col].apply(lambda v: _normalize_snapshot(v)[1])
            mask = old_keys == compare_key
            removed = int(mask.sum())
            old = old.loc[~mask]
        else:
            # Fix: without the snapshot column there is nothing to replace,
            # so this silently degrades to a blind append. Say so.
            warnings.warn(
                f"[{sheet_name}] rerun_mode='replace_snapshot' but column "
                f"'{snapshot_col}' is absent from the existing master. "
                f"Rows were appended without replacement; re-running today "
                f"will duplicate them.",
                stacklevel=2,
            )

    combined = pd.concat([old, new], ignore_index=True) if not old.empty else new
    cols = [snapshot_col] + [c for c in combined.columns if c != snapshot_col]
    combined = combined[cols]

    if keep_snapshots and snapshot_col in combined.columns:
        keys = combined[snapshot_col].apply(lambda v: _normalize_snapshot(v)[1])
        recent = sorted(keys.unique())[-keep_snapshots:]
        combined = combined.loc[keys.isin(recent)]

    _write_master(combined, master_path, sheet_name, keep_backup, backup_dir, backup_keep)

    if verbose:
        note = f", {intra} intra-batch dupes dropped" if intra else ""
        print(
            f"[{sheet_name}] snapshot {compare_key}: "
            f"+{added} rows, -{removed} replaced{note}, "
            f"master now {len(combined)} rows."
        )
    return combined


def _write_master(combined, master_path, sheet_name, keep_backup, backup_dir=None, backup_keep=10):
    os.makedirs(os.path.dirname(os.path.abspath(master_path)) or ".", exist_ok=True)

    if os.path.exists(master_path):
        if backup_dir:
            _rotate_backup(master_path, backup_dir, backup_keep)
        elif keep_backup:
            shutil.copy2(master_path, master_path + ".bak")
        wb = load_workbook(master_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws.append(list(combined.columns))
    date_col_idxs = {
        i for i, c in enumerate(combined.columns)
        if combined[c].map(lambda v: isinstance(v, (datetime, date))).any()
    }
    for row in combined.itertuples(index=False, name=None):
        ws.append([_clean(v) for v in row])

    header_fill = PatternFill("solid", start_color="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # Apply a real date number format to any column holding date values so
    # Excel and downstream BI tools recognize it as a Date, not text.
    for idx in date_col_idxs:
        col_letter = get_column_letter(idx + 1)
        for r in range(2, ws.max_row + 1):
            ws[f"{col_letter}{r}"].number_format = "yyyy-mm-dd"

    for idx, col in enumerate(combined.columns, start=1):
        sample = combined[col].head(200).map(lambda v: len(str(v)) if pd.notna(v) else 0)
        width = min(max(int(sample.max() if len(sample) else 10), len(str(col))) + 2, 45)
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Fix: an Excel table needs at least one body row. Skip the table on an
    # empty result rather than writing a workbook Excel will flag as corrupt.
    if len(combined) > 0:
        last_col = get_column_letter(len(combined.columns))
        ref = f"A1:{last_col}{len(combined) + 1}"
        tname = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in sheet_name)
        table = Table(displayName=tname, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    # Write to a temp file in the same directory, then atomically replace.
    # A crash mid-write leaves the previous master intact rather than a
    # truncated file the next run would read as the master.
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(os.path.abspath(master_path)))
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, master_path)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


def _rotate_backup(master_path, backup_dir, keep):
    os.makedirs(backup_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(master_path))[0]
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(master_path, os.path.join(backup_dir, f"{base}_{stamp}.xlsx"))
    backups = sorted(
        f for f in os.listdir(backup_dir)
        if f.startswith(base + "_") and f.endswith(".xlsx")
    )
    for stale in (backups[:-keep] if keep else []):
        os.remove(os.path.join(backup_dir, stale))


def _clean(v):
    # Fix: pd.isna returns an array for list-like input, which raises on a
    # truth test. Guard by type rather than relying on a conditional
    # expression that only excluded list and dict.
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, (list, dict, tuple, set)):
        return str(v)
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        return v
    return v
